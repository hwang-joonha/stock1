#!/usr/bin/env python3
"""xlsx에 박힌 수식을 셀 참조를 따라가며 평가한다.

G7이 검사하려는 것은 **엔진의 astToExcel이 모델 수식을 Excel 셀 수식으로
옮기면서 뜻을 보존하는가**이다. 그러려면 옮겨진 셀 수식을 원래 계산과는
독립된 경로로 다시 계산해봐야 한다.

여기서 하는 일이 그것이다. openpyxl로 셀 수식 문자열을 읽어 직접 파싱하고,
셀 참조를 실제 행·열로 따라가며 값을 만든다. 행 배치가 어긋났거나, PREV가
엉뚱한 열을 가리키거나, IF 인자 순서가 뒤집혔다면 여기서 값이 달라진다.

LibreOffice로 재계산하는 편이 더 독립적이지만 이 컨테이너의 설치가
xlsx를 열지 못한다(최소 파일도 실패). 그래서 직접 평가한다.
"""
from __future__ import annotations

import re

from openpyxl.utils import column_index_from_string

# 토큰: 문자열, 시트참조 셀, 셀, 숫자, 함수명, 연산자, 괄호, 쉼표
_TOKEN = re.compile(r"""
      (?P<ws>\s+)
    | (?P<str>"(?:[^"]|"")*")
    | (?P<ref>(?:'[^']+'|[A-Za-z_][A-Za-z0-9_ .]*)!\$?[A-Z]{1,3}\$?\d+)
    | (?P<cell>\$?[A-Z]{1,3}\$?\d+)
    | (?P<num>\d+(?:\.\d+)?(?:[eE][-+]?\d+)?)
    | (?P<fn>[A-Za-z_][A-Za-z0-9_.]*)
    | (?P<op><=|>=|<>|[-+*/^&<>=])
    | (?P<par>[()])
    | (?P<comma>[,;])
""", re.X)


class FormulaError(Exception):
    pass


def _tokenize(src: str) -> list[tuple[str, str]]:
    out, i = [], 0
    while i < len(src):
        m = _TOKEN.match(src, i)
        if not m:
            raise FormulaError(f"토큰 인식 실패: {src[i:i + 20]!r}")
        i = m.end()
        kind = m.lastgroup
        if kind == "ws":
            continue
        out.append((kind, m.group()))
    return out


class Evaluator:
    """워크북 하나에 대한 수식 평가기. 셀 값은 메모이즈한다."""

    def __init__(self, wb, default_sheet: str):
        self.wb = wb
        self.default = default_sheet
        self._cache: dict[tuple[str, str], float] = {}
        self._stack: set[tuple[str, str]] = set()

    # ── 셀 값 ────────────────────────────────────────────────
    def cell(self, sheet: str, addr: str) -> float:
        key = (sheet, addr.replace("$", ""))
        if key in self._cache:
            return self._cache[key]
        if key in self._stack:
            raise FormulaError(f"셀 순환 참조: {sheet}!{addr}")
        self._stack.add(key)
        try:
            raw = self.wb[sheet][key[1]].value
            if isinstance(raw, str) and raw.startswith("="):
                v = self.evaluate(raw[1:], sheet)
            elif isinstance(raw, (int, float)):
                v = float(raw)
            else:
                v = 0.0          # 빈 셀은 0. Excel과 같은 취급이다.
            self._cache[key] = v
            return v
        finally:
            self._stack.discard(key)

    # ── 파서 ────────────────────────────────────────────────
    # 파서 상태를 인스턴스에 두면 안 된다. 셀 참조를 따라가다 보면 evaluate가
    # 자기 자신을 재귀 호출하고, 그때 바깥 호출의 토큰 위치가 덮여버린다.
    # 호출마다 별도 파서를 만든다.
    def evaluate(self, expr: str, sheet: str | None = None) -> float:
        return _Parser(self, sheet or self.default, _tokenize(expr)).parse()


class _Parser:
    """수식 하나를 훑는 재귀 하강 파서. 호출 1회당 1개."""

    def __init__(self, ev: "Evaluator", sheet: str, toks: list[tuple[str, str]]):
        self.ev = ev
        self.sheet = sheet
        self.toks = toks
        self.pos = 0

    def parse(self) -> float:
        v = self._comparison()
        if self.pos != len(self.toks):
            raise FormulaError(f"남은 토큰: {self.toks[self.pos:]}")
        return v

    def _peek(self):
        return self.toks[self.pos] if self.pos < len(self.toks) else (None, None)

    def _eat(self, val=None):
        kind, tok = self._peek()
        if kind is None or (val is not None and tok != val):
            raise FormulaError(f"기대: {val}, 실제: {tok}")
        self.pos += 1
        return tok

    def _comparison(self) -> float:
        left = self._sum()
        kind, tok = self._peek()
        if kind == "op" and tok in ("=", "<>", "<", ">", "<=", ">="):
            self.pos += 1
            right = self._sum()
            return float({
                "=": left == right, "<>": left != right,
                "<": left < right, ">": left > right,
                "<=": left <= right, ">=": left >= right,
            }[tok])
        return left

    def _sum(self) -> float:
        v = self._term()
        while True:
            kind, tok = self._peek()
            if kind == "op" and tok in "+-":
                self.pos += 1
                r = self._term()
                v = v + r if tok == "+" else v - r
            else:
                return v

    def _term(self) -> float:
        v = self._unary()
        while True:
            kind, tok = self._peek()
            if kind == "op" and tok in "*/":
                self.pos += 1
                r = self._unary()
                if tok == "*":
                    v *= r
                else:
                    # 엔진의 나눗셈 규약과 같다 — 0으로 나누면 오류가 아니라 0.
                    v = 0.0 if r == 0 else v / r
            else:
                return v

    def _unary(self) -> float:
        kind, tok = self._peek()
        if kind == "op" and tok in "+-":
            self.pos += 1
            v = self._unary()
            return -v if tok == "-" else v
        return self._atom()

    def _atom(self) -> float:
        kind, tok = self._peek()
        if kind == "par" and tok == "(":
            self.pos += 1
            v = self._comparison()
            self._eat(")")
            return v
        if kind == "num":
            self.pos += 1
            return float(tok)
        if kind == "cell":
            self.pos += 1
            return self.ev.cell(self.sheet, tok)
        if kind == "ref":
            self.pos += 1
            sh, addr = tok.split("!", 1)
            return self.ev.cell(sh.strip("'"), addr)
        if kind == "fn":
            self.pos += 1
            name = tok.upper()
            if self._peek()[1] != "(":
                raise FormulaError(f"함수 괄호 없음: {name}")
            self.pos += 1
            args = []
            if self._peek()[1] != ")":
                args.append(self._comparison())
                while self._peek()[0] == "comma":
                    self.pos += 1
                    args.append(self._comparison())
            self._eat(")")
            return self._call(name, args)
        raise FormulaError(f"해석 불가 토큰: {tok!r}")

    def _call(self, name: str, args: list[float]) -> float:
        if name == "SUM":
            return sum(args)
        if name == "MIN":
            return min(args)
        if name == "MAX":
            return max(args)
        if name == "AVERAGE":
            return sum(args) / len(args) if args else 0.0
        if name == "IF":
            if len(args) != 3:
                raise FormulaError("IF는 인자 3개")
            return args[1] if args[0] else args[2]
        if name == "ROUND":
            return round(args[0], int(args[1]) if len(args) > 1 else 0)
        if name == "ABS":
            return abs(args[0])
        raise FormulaError(f"미지원 함수: {name}")


def sheet_values(wb, sheet: str, row: int, col0: int, count: int) -> list[float]:
    """한 행의 연속된 셀을 평가해 값 배열로 돌려준다."""
    ev = Evaluator(wb, sheet)
    from openpyxl.utils import get_column_letter
    return [ev.cell(sheet, f"{get_column_letter(col0 + i)}{row}")
            for i in range(count)]


__all__ = ["Evaluator", "FormulaError", "sheet_values", "column_index_from_string"]
