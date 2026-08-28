#!/usr/bin/env python3
"""model.html → model.xlsx. 수식이 살아있는 Excel 모델을 만든다.

**수식 변환의 정본은 JS 엔진(astToExcel)이다.** 이 스크립트는 harness가
JS에서 받아온 셀 수식을 배치만 한다. 파이썬에 같은 변환기를 두면 변환기가
둘이 되고, 둘이 갈라지는 순간 G7 대사가 자기 자신을 검사하게 된다.

사용:
    python3 tools/build_excel.py companies/samsung-em/model.html
"""
from __future__ import annotations

import argparse
import datetime
import hashlib
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

from harness import excel_payload  # noqa: E402

# 금융모델링 관행 — 셀 색이 곧 셀의 성격이다 (framework/excel_spec.md §2).
INPUT_FONT = Font(color="1F4E9C", name="Calibri", size=10)          # 파랑 = 하드코딩
CALC_FONT = Font(color="1C1C1C", name="Calibri", size=10)           # 검정 = 수식
LABEL_FONT = Font(color="1C1C1C", name="Calibri", size=10)
BOLD = Font(bold=True, color="1C1C1C", name="Calibri", size=10)
TITLE = Font(bold=True, size=13, color="1E2185", name="Calibri")
HEAD_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
MUTED = Font(color="6B7280", name="Calibri", size=9)

HEAD_FILL = PatternFill("solid", fgColor="1E2185")
HIST_FILL = PatternFill("solid", fgColor="EFEFF2")   # 회색 = 실적 구간
FC_FILL = PatternFill("solid", fgColor="EDF3FF")     # 옅은 파랑 = 추정 구간
ROOT_FILL = PatternFill("solid", fgColor="DFE8FF")

THIN = Side(style="thin", color="D8DCE3")
BOX = Border(bottom=THIN)

DATA_COL0 = 4          # D열부터 연도. astToExcel의 _colLetter(3+yr)와 같은 약속.

# 문서 속성 타임스탬프를 고정한다. 빌드를 재현 가능하게 만들기 위한 것이고,
# 실제 빌드 정보는 Metadata 시트에 남는다.
_EPOCH = datetime.datetime(2000, 1, 1)
SHEETS = ["Index", "Control", "Assumptions", "Model",
          "Formula Audit", "Structure", "Checks", "Metadata"]


def _year_headers(ws, payload: dict, row_band: int, row_year: int,
                  first_col: int = DATA_COL0) -> None:
    """실적/추정 밴드와 연도 머리말. HIST_N 하나로 결정한다 — 하드코딩 금지."""
    yrs, hist_n = payload["YRS"], payload["HIST_N"]
    for i, y in enumerate(yrs):
        c = first_col + i
        band = ws.cell(row=row_band, column=c)
        band.value = "실적" if i < hist_n else "추정"
        band.font = MUTED
        band.alignment = Alignment(horizontal="center")
        band.fill = HIST_FILL if i < hist_n else FC_FILL
        h = ws.cell(row=row_year, column=c)
        h.value = y + ("A" if i < hist_n else "E")
        h.font = HEAD_FONT
        h.fill = HEAD_FILL
        h.alignment = Alignment(horizontal="center")


def _widths(ws, spec: dict) -> None:
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


def _sheet_model(wb: Workbook, payload: dict) -> None:
    """본체. 계정 트리 × 연도, 수식이 그대로 박힌다."""
    ws = wb["Model"]
    yrs, hist_n = payload["YRS"], payload["HIST_N"]
    meta = payload.get("meta", {})
    units = payload.get("units", {})

    ws["A1"] = meta.get("title", "추정 모델")
    ws["A1"].font = TITLE
    ws["A2"] = f"단위 {units.get('money', '')} · 실적 {hist_n}개년 + 추정 {len(yrs) - hist_n}개년"
    ws["A2"].font = MUTED
    for a, t in (("A4", "계정"), ("B4", "노드 id"), ("C4", "단위")):
        ws[a] = t
        ws[a].font = HEAD_FONT
        ws[a].fill = HEAD_FILL
    _year_headers(ws, payload, row_band=3, row_year=4)

    for node_id in payload["order"]:
        n = payload["nodes"][node_id]
        r = payload["rowMap"][node_id]
        lab = ws.cell(row=r, column=1, value=n["label"])
        lab.alignment = Alignment(indent=min(n["depth"], 8))
        lab.font = BOLD if n["depth"] <= 1 else LABEL_FONT
        ws.cell(row=r, column=2, value=node_id).font = MUTED
        ws.cell(row=r, column=3, value=n["unit"]).font = MUTED

        formulas = payload["formulas"].get(node_id)
        for i in range(len(yrs)):
            c = ws.cell(row=r, column=DATA_COL0 + i)
            if formulas and formulas[i]:
                c.value = formulas[i]
                c.font = CALC_FONT
            else:
                c.value = n["values"][i] if i < len(n["values"]) else 0
                c.font = INPUT_FONT      # 파랑 = 손으로 넣은 값
            c.number_format = n["numfmt"]
            if i < hist_n:
                c.fill = HIST_FILL
            if n["depth"] == 0:
                c.fill = ROOT_FILL
            c.border = BOX

    _widths(ws, {"A": 30, "B": 22, "C": 12})
    for i in range(len(yrs)):
        ws.column_dimensions[get_column_letter(DATA_COL0 + i)].width = 14
    ws.freeze_panes = ws.cell(row=5, column=DATA_COL0)


def _sheet_assumptions(wb: Workbook, payload: dict) -> None:
    """입력 노드 일람. 근거(desc)를 함께 실어 가정의 출처가 파일 안에 남게 한다."""
    ws = wb["Assumptions"]
    ws["A1"] = "가정변수 일람"
    ws["A1"].font = TITLE
    ws["A2"] = "파랑 = 하드코딩 입력값. 근거 열의 대괄호 태그가 객관/주관 구분이다."
    ws["A2"].font = MUTED
    for a, t in (("A4", "계정"), ("B4", "노드 id"), ("C4", "단위")):
        ws[a] = t
        ws[a].font = HEAD_FONT
        ws[a].fill = HEAD_FILL
    _year_headers(ws, payload, row_band=3, row_year=4)
    desc_col = DATA_COL0 + len(payload["YRS"])
    d4 = ws.cell(row=4, column=desc_col, value="근거")
    d4.font = HEAD_FONT
    d4.fill = HEAD_FILL

    r = 5
    for node_id in payload["order"]:
        n = payload["nodes"][node_id]
        if n["type"] != "input":
            continue
        ws.cell(row=r, column=1, value=n["label"]).font = LABEL_FONT
        ws.cell(row=r, column=2, value=node_id).font = MUTED
        ws.cell(row=r, column=3, value=n["unit"]).font = MUTED
        for i in range(len(payload["YRS"])):
            c = ws.cell(row=r, column=DATA_COL0 + i,
                        value=n["values"][i] if i < len(n["values"]) else 0)
            c.font = INPUT_FONT
            c.number_format = n["numfmt"]
            if i < payload["HIST_N"]:
                c.fill = HIST_FILL
        dc = ws.cell(row=r, column=desc_col, value=n["desc"])
        dc.alignment = Alignment(wrap_text=True, vertical="top")
        dc.font = MUTED
        r += 1

    _widths(ws, {"A": 30, "B": 22, "C": 12,
                 get_column_letter(desc_col): 90})
    for i in range(len(payload["YRS"])):
        ws.column_dimensions[get_column_letter(DATA_COL0 + i)].width = 13
    ws.freeze_panes = ws.cell(row=5, column=DATA_COL0)


def _sheet_formula_audit(wb: Workbook, payload: dict) -> None:
    """모델 문법의 원식과 첫 추정연도 셀 수식을 나란히 둔다.

    수식이 어떻게 옮겨졌는지 사람이 눈으로 대조할 수 있어야 한다.
    """
    ws = wb["Formula Audit"]
    ws["A1"] = "수식 감사"
    ws["A1"].font = TITLE
    ws["A2"] = "모델 원식 ↔ Excel 셀 수식 대조. 변환은 엔진의 astToExcel이 담당한다."
    ws["A2"].font = MUTED
    heads = ["노드 id", "계정", "행", "모델 원식", "첫 추정연도 셀 수식"]
    for i, t in enumerate(heads, start=1):
        c = ws.cell(row=4, column=i, value=t)
        c.font = HEAD_FONT
        c.fill = HEAD_FILL

    r, fc0 = 5, payload["HIST_N"]
    for node_id in payload["order"]:
        n = payload["nodes"][node_id]
        if n["type"] != "computed":
            continue
        fs = payload["formulas"].get(node_id) or []
        ws.cell(row=r, column=1, value=node_id).font = MUTED
        ws.cell(row=r, column=2, value=n["label"]).font = LABEL_FONT
        ws.cell(row=r, column=3, value=payload["rowMap"][node_id]).font = MUTED
        f1 = ws.cell(row=r, column=4, value=n["formula"])
        f1.alignment = Alignment(wrap_text=True, vertical="top")
        f1.font = LABEL_FONT
        # 앞에 작은따옴표를 붙여 텍스트로 둔다. 수식으로 들어가면 여기서 계산된다.
        f2 = ws.cell(row=r, column=5,
                     value=("'" + fs[fc0]) if fc0 < len(fs) and fs[fc0] else "")
        f2.alignment = Alignment(wrap_text=True, vertical="top")
        f2.font = MUTED
        r += 1
    _widths(ws, {"A": 22, "B": 26, "C": 6, "D": 60, "E": 60})


def _sheet_structure(wb: Workbook, payload: dict) -> None:
    ws = wb["Structure"]
    ws["A1"] = "트리 구조"
    ws["A1"].font = TITLE
    heads = ["노드 id", "계정", "깊이", "유형", "합계형", "행", "단위"]
    for i, t in enumerate(heads, start=1):
        c = ws.cell(row=3, column=i, value=t)
        c.font = HEAD_FONT
        c.fill = HEAD_FILL
    for r, node_id in enumerate(payload["order"], start=4):
        n = payload["nodes"][node_id]
        ws.cell(row=r, column=1, value=node_id).font = MUTED
        lab = ws.cell(row=r, column=2, value=n["label"])
        lab.alignment = Alignment(indent=min(n["depth"], 8))
        ws.cell(row=r, column=3, value=n["depth"])
        ws.cell(row=r, column=4, value=n["type"])
        ws.cell(row=r, column=5, value="Y" if n["isSum"] else "")
        ws.cell(row=r, column=6, value=payload["rowMap"][node_id])
        ws.cell(row=r, column=7, value=n["unit"])
    _widths(ws, {"A": 22, "B": 30, "C": 7, "D": 12, "E": 9, "F": 7, "G": 12})


def _sheet_checks(wb: Workbook, payload: dict) -> None:
    """정합 검증을 Excel 수식으로 심는다.

    파이썬이 계산한 결과를 적어두면 파일을 열어 값을 바꿔도 검증이 따라오지
    않는다. 시트 안에서 살아있는 수식이어야 검증이 실제로 작동한다.
    """
    ws = wb["Checks"]
    ws["A1"] = "검증"
    ws["A1"].font = TITLE
    ws["A2"] = "합계형 노드의 부모값과 자식 합의 차이. 전 칸이 0이어야 한다."
    ws["A2"].font = MUTED
    for a, t in (("A4", "계정"), ("B4", "검사")):
        ws[a] = t
        ws[a].font = HEAD_FONT
        ws[a].fill = HEAD_FILL
    _year_headers(ws, payload, row_band=3, row_year=4)

    # order는 깊이 우선 순회 결과다. depth로 스택을 되감으면 부모-자식이 복원된다.
    kids: dict[str, list[str]] = {}
    stack: list[str] = []
    for node_id in payload["order"]:
        d = payload["nodes"][node_id]["depth"]
        stack = stack[:d]
        if stack:
            kids.setdefault(stack[-1], []).append(node_id)
        stack.append(node_id)

    r = 5
    for node_id in payload["order"]:
        n = payload["nodes"][node_id]
        if not n["isSum"] or node_id not in kids:
            continue
        ws.cell(row=r, column=1, value=n["label"]).font = LABEL_FONT
        ws.cell(row=r, column=2, value="부모 − 자식합").font = MUTED
        prow = payload["rowMap"][node_id]
        for i in range(len(payload["YRS"])):
            col = get_column_letter(DATA_COL0 + i)
            terms = "+".join(f"Model!{col}{payload['rowMap'][k]}" for k in kids[node_id])
            c = ws.cell(row=r, column=DATA_COL0 + i,
                        value=f"=ROUND(Model!{col}{prow}-({terms}),6)")
            c.font = CALC_FONT
            c.number_format = "0.000000"
        r += 1

    ws.cell(row=r + 1, column=1, value="합계형 노드 수").font = BOLD
    ws.cell(row=r + 1, column=2, value=len([k for k in kids
                                            if payload["nodes"][k]["isSum"]]))
    _widths(ws, {"A": 30, "B": 16})
    for i in range(len(payload["YRS"])):
        ws.column_dimensions[get_column_letter(DATA_COL0 + i)].width = 14


def _sheet_control(wb: Workbook, payload: dict) -> None:
    ws = wb["Control"]
    ws["A1"] = "컨트롤"
    ws["A1"].font = TITLE
    ws["A2"] = ("가정을 바꾸려면 Assumptions 시트가 아니라 Model 시트의 파랑 셀을 "
                "고친다. Model이 계산의 단일 출처다.")
    ws["A2"].font = MUTED
    rows = [
        ("모델", payload.get("meta", {}).get("title", "")),
        ("통화 단위", payload.get("units", {}).get("money", "")),
        ("범위", payload.get("units", {}).get("scope", "")),
        ("실적 연도 수 (HIST_N)", payload["HIST_N"]),
        ("추정 연도 수", len(payload["YRS"]) - payload["HIST_N"]),
        ("첫 연도", payload["YRS"][0]),
        ("마지막 연도", payload["YRS"][-1]),
    ]
    for i, (k, v) in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=k).font = BOLD
        ws.cell(row=i, column=2, value=v)
    _widths(ws, {"A": 26, "B": 60})


def _sheet_index(wb: Workbook, payload: dict) -> None:
    ws = wb["Index"]
    ws["A1"] = payload.get("meta", {}).get("title", "추정 모델")
    ws["A1"].font = TITLE
    ws["A2"] = "tools/build_excel.py가 model.html에서 생성. 직접 고치지 말 것 — 다음 빌드에 덮인다."
    ws["A2"].font = MUTED
    guide = {
        "Index": "이 시트",
        "Control": "모델 개요·전역 설정",
        "Assumptions": "가정변수 일람과 근거",
        "Model": "계정 트리 × 연도. 수식이 살아있는 본체",
        "Formula Audit": "모델 원식 ↔ 셀 수식 대조",
        "Structure": "트리 구조·깊이·행 배치",
        "Checks": "부모 = 자식 합 정합 검증",
        "Metadata": "빌드 정보",
    }
    ws.cell(row=4, column=1, value="시트").font = HEAD_FONT
    ws.cell(row=4, column=1).fill = HEAD_FILL
    ws.cell(row=4, column=2, value="내용").font = HEAD_FONT
    ws.cell(row=4, column=2).fill = HEAD_FILL
    for i, (name, desc) in enumerate(guide.items(), start=5):
        ws.cell(row=i, column=1, value=name).font = BOLD
        ws.cell(row=i, column=2, value=desc).font = LABEL_FONT
    ws.cell(row=len(guide) + 6, column=1, value="셀 색 규약").font = BOLD
    for i, (color, meaning) in enumerate([
        ("파랑 글자", "하드코딩 입력값 — 여기만 고친다"),
        ("검정 글자", "수식"),
        ("회색 배경", "실적(Historical) 구간 — 공시 확정값"),
    ], start=len(guide) + 7):
        ws.cell(row=i, column=1, value=color).font = MUTED
        ws.cell(row=i, column=2, value=meaning).font = MUTED
    _widths(ws, {"A": 20, "B": 70})


def _sheet_metadata(wb: Workbook, payload: dict, src: str) -> None:
    """빌드 정보.

    빌드 시각은 넣지 않는다. 넣으면 같은 입력에서 매번 다른 파일이 나와
    "재빌드 후 diff 없음"이라는 검증이 성립하지 않는다. 대신 소스의 내용
    해시를 남긴다 — 어느 model.html에서 나왔는지가 시각보다 유용하기도 하다.
    """
    ws = wb["Metadata"]
    ws["A1"] = "빌드 정보"
    ws["A1"].font = TITLE
    with open(src, "rb") as fh:
        src_hash = hashlib.sha256(fh.read()).hexdigest()[:16]
    rows = [
        ("소스", os.path.basename(src)),
        ("소스 해시 (sha256 앞 16자리)", src_hash),
        ("빌더", "tools/build_excel.py"),
        ("수식 변환", "엔진 astToExcel (JS) — 파이썬은 배치만 한다"),
        ("노드 수", len(payload["order"])),
        ("입력 노드 수", sum(1 for k in payload["order"]
                        if payload["nodes"][k]["type"] == "input")),
        ("연도 수", len(payload["YRS"])),
        ("HIST_N", payload["HIST_N"]),
        ("모델 id", payload.get("meta", {}).get("modelId", "")),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=k).font = BOLD
        ws.cell(row=i, column=2, value=v)
    _widths(ws, {"A": 22, "B": 60})


def build(html_path: str, out_path: str) -> dict:
    payload = excel_payload(html_path)
    wb = Workbook()
    wb.remove(wb.active)
    for name in SHEETS:
        wb.create_sheet(name)
    _sheet_index(wb, payload)
    _sheet_control(wb, payload)
    _sheet_assumptions(wb, payload)
    _sheet_model(wb, payload)
    _sheet_formula_audit(wb, payload)
    _sheet_structure(wb, payload)
    _sheet_checks(wb, payload)
    _sheet_metadata(wb, payload, html_path)
    wb.properties.created = wb.properties.modified = _EPOCH
    wb.properties.creator = "tools/build_excel.py"
    wb.save(out_path)
    _freeze_timestamps(out_path)
    return payload


def _freeze_timestamps(path: str) -> None:
    """저장 시각을 문서 속성에서 걷어낸다.

    openpyxl은 save() 안에서 modified를 현재 시각으로 다시 덮는다. 그대로 두면
    같은 입력에서 매번 다른 바이트가 나오고, "재빌드 후 git diff 없음"이라는
    검증이 성립하지 않는다. 저장된 zip의 core.xml만 다시 쓴다.
    """
    import re
    import shutil
    import zipfile

    stamp = _EPOCH.strftime("%Y-%m-%dT%H:%M:%SZ")
    tmp = path + ".tmp"
    with zipfile.ZipFile(path) as src, \
            zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as dst:
        for item in src.infolist():
            data = src.read(item.filename)
            if item.filename == "docProps/core.xml":
                text = data.decode("utf-8")
                # 그룹 참조는 반드시 \g<n> 형태로. \1 뒤에 숫자가 오면
                # 8진 이스케이프로 읽혀 여는 태그가 통째로 사라진다.
                text = re.sub(r"(<dcterms:(?:created|modified)[^>]*>)[^<]*(</)",
                              rf"\g<1>{stamp}\g<2>", text)
                data = text.encode("utf-8")
            info = zipfile.ZipInfo(item.filename, date_time=(1980, 1, 1, 0, 0, 0))
            info.compress_type = zipfile.ZIP_DEFLATED
            info.external_attr = item.external_attr
            dst.writestr(info, data)
    shutil.move(tmp, path)


def main(argv: list[str]) -> int:
    ap = argparse.ArgumentParser(description="model.html → model.xlsx")
    ap.add_argument("model", help="model.html 경로")
    ap.add_argument("-o", "--out", help="출력 경로 (기본: model.xlsx)")
    args = ap.parse_args(argv)
    out = args.out or os.path.splitext(args.model)[0] + ".xlsx"
    payload = build(args.model, out)
    print(f"{out} 빌드 완료 — {len(payload['order'])}개 노드 × "
          f"{len(payload['YRS'])}개년, 시트 {len(SHEETS)}개")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
